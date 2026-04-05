"""
Excel 配置读取工具
基于 openpyxl，将策划 Excel 配置转为 AI 可读的文本

用法示例:
  # 搜索包含关键词的配置（搜索文件名、sheet 名和单元格数据）
  python tools/excel_reader.py --search "活动名称"

  # 在指定文件中搜索（更快）
  python tools/excel_reader.py --search "10001" --file item.xlsx

  # 读取指定 Excel 文件的指定 sheet
  python tools/excel_reader.py --file "活动配置.xlsx" --sheet "Sheet1"

  # 按某列过滤
  python tools/excel_reader.py --file "活动配置.xlsx" --sheet "Sheet1" --filter "活动名称=新手任务"

  # 列出某个 Excel 文件的所有 sheet 名
  python tools/excel_reader.py --file "活动配置.xlsx" --list-sheets

  # 以 markdown 格式输出
  python tools/excel_reader.py --file "活动配置.xlsx" --sheet "Sheet1" --format markdown
"""

import argparse
import sys
from pathlib import Path

import openpyxl

from utils import load_config, format_json, format_markdown_table, print_output, get_repo_path


def get_config_path(config):
    """获取配表仓库路径"""
    config_path = get_repo_path(config, "config")
    if not config_path:
        print("Error: code.repos.config not configured in config.yaml", file=sys.stderr)
        sys.exit(1)
    path = Path(config_path)
    if not path.exists():
        print(f"Error: config path does not exist: {path}", file=sys.stderr)
        sys.exit(1)
    return path


def _cell_to_str(val):
    """将单元格值转为字符串用于搜索匹配"""
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val)


def search_excel_files(config_path, keyword, target_file=None):
    """
    在配置目录中搜索包含关键词的 Excel 文件、sheet 名和单元格数据

    Args:
        config_path: 配表仓库路径
        keyword: 搜索关键词
        target_file: 限定搜索范围到指定文件（可选，加速搜索）
    Returns:
        list: 匹配结果列表
    """
    results = []
    keyword_lower = keyword.lower()

    if target_file:
        files = [resolve_file_path(config_path, target_file)]
    else:
        files = sorted(Path(config_path).rglob("*.xlsx"))

    for f in files:
        if f.name.startswith("~$"):
            continue

        rel_path = str(f.relative_to(config_path))

        if not target_file and keyword_lower in f.stem.lower():
            results.append({
                "file": rel_path,
                "match_type": "filename",
                "match_value": f.stem,
            })

        try:
            wb = openpyxl.load_workbook(f, read_only=True)
            for sheet_name in wb.sheetnames:
                if keyword_lower in sheet_name.lower():
                    results.append({
                        "file": rel_path,
                        "sheet": sheet_name,
                        "match_type": "sheet_name",
                        "match_value": sheet_name,
                    })

                # 搜索单元格数据
                ws = wb[sheet_name]
                headers = []
                for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                    headers = [str(c) if c is not None else "" for c in row]
                    break
                if not headers:
                    continue

                for row in ws.iter_rows(min_row=2, values_only=True):
                    first_val = _cell_to_str(row[0] if row else "")
                    if first_val.startswith("##"):
                        continue
                    for i, cell in enumerate(row):
                        cell_str = _cell_to_str(cell)
                        if cell_str and keyword_lower in cell_str.lower():
                            col_name = headers[i] if i < len(headers) else f"col_{i}"
                            key_val = _cell_to_str(row[1] if len(row) > 1 else row[0])
                            results.append({
                                "file": rel_path,
                                "sheet": sheet_name,
                                "match_type": "cell_value",
                                "column": col_name,
                                "row_key": key_val,
                                "match_value": cell_str[:100],
                            })
                            break  # 每行只报一次

            wb.close()
        except Exception as e:
            print(f"Warning: cannot read {rel_path}: {e}", file=sys.stderr)
            continue

    return results


def list_sheets(filepath):
    """
    列出 Excel 文件的所有 sheet 名

    Args:
        filepath: Excel 文件路径
    Returns:
        list: sheet 名列表
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    return sheets


def read_sheet(filepath, sheet_name, filter_expr=None, limit=50):
    """
    读取 Excel sheet 并返回结构化数据

    Args:
        filepath: Excel 文件路径
        sheet_name: sheet 名称
        filter_expr: 过滤表达式 "列名=值"
        limit: 返回行数上限
    Returns:
        dict: {headers, rows, total_shown, total_matched}
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)

    if sheet_name not in wb.sheetnames:
        available = wb.sheetnames
        wb.close()
        print(f"Error: sheet '{sheet_name}' not found. Available sheets: {available}", file=sys.stderr)
        sys.exit(1)

    ws = wb[sheet_name]

    # 读取表头（第一行）
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(cell) if cell is not None else "" for cell in row]
        break

    if not headers:
        wb.close()
        return {"headers": [], "rows": [], "total_shown": 0, "total_matched": 0}

    # 解析过滤条件
    filter_key = None
    filter_val = None
    if filter_expr and "=" in filter_expr:
        filter_key, filter_val = filter_expr.split("=", 1)

    # 读取数据行（跳过 ## 开头的元数据行）
    rows = []
    meta_rows = []
    total_matched = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for i, val in enumerate(row):
            if i < len(headers):
                record[headers[i]] = val

        # 检测 ## 元数据行（第一列以 ## 开头，如 ##type, ##group, ##）
        first_val = str(record.get(headers[0], "") or "")
        if first_val.startswith("##"):
            meta_rows.append(record)
            continue

        # 应用过滤
        if filter_key and filter_val:
            cell_val = str(record.get(filter_key, ""))
            if filter_val.lower() not in cell_val.lower():
                continue

        total_matched += 1
        if len(rows) < limit:
            rows.append(record)

    wb.close()

    # 从 ## 注释行中提取字段描述
    field_descriptions = {}
    for meta in meta_rows:
        first_val = str(meta.get(headers[0], "") or "")
        if first_val == "##":
            for h in headers:
                desc = meta.get(h)
                if desc and str(desc).strip() and not str(desc).startswith("##"):
                    field_descriptions[h] = str(desc).replace("\n", " ")
            break

    return {
        "headers": headers,
        "field_descriptions": field_descriptions,
        "rows": rows,
        "total_shown": len(rows),
        "total_matched": total_matched,
    }


def resolve_file_path(config_path, file_arg):
    """
    解析文件路径，支持相对路径和绝对路径

    Args:
        config_path: 配表仓库根路径
        file_arg: 用户传入的文件路径
    Returns:
        Path: 完整文件路径
    """
    # 先尝试相对于配表目录
    full_path = Path(config_path) / file_arg
    if full_path.exists():
        return full_path

    # 再尝试绝对路径
    abs_path = Path(file_arg)
    if abs_path.exists():
        return abs_path

    # 尝试模糊匹配（不带目录前缀搜索文件名）
    filename = Path(file_arg).name
    matches = list(Path(config_path).rglob(filename))
    if len(matches) == 1:
        return matches[0]
    elif len(matches) > 1:
        print(f"Error: multiple files matching '{filename}':", file=sys.stderr)
        for m in matches:
            print(f"  - {m.relative_to(config_path)}", file=sys.stderr)
        sys.exit(1)

    print(f"Error: file not found: {file_arg} (searched in {config_path})", file=sys.stderr)
    sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description="Excel config reader for BugHunter")
    parser.add_argument("--config", help="Path to config.yaml")

    # 操作模式
    parser.add_argument("--search", help="Search Excel files by keyword (searches filename and sheet names)")
    parser.add_argument("--file", help="Excel file path (relative to config_path or absolute)")
    parser.add_argument("--sheet", help="Sheet name to read")
    parser.add_argument("--list-sheets", action="store_true", help="List all sheets in the Excel file")

    # 过滤和输出
    parser.add_argument("--filter", help="Filter rows by column value: 'column_name=value'")
    parser.add_argument("--limit", type=int, default=50, help="Max rows to output (default: 50)")
    parser.add_argument("--format", choices=["json", "markdown"], default="json", help="Output format (default: json)")

    args = parser.parse_args()
    config = load_config(args.config)
    config_path = get_config_path(config)

    # 搜索模式
    if args.search:
        target = args.file if args.file else None
        scope = args.file or str(config_path)
        print(f"Searching for '{args.search}' in {scope}", file=sys.stderr)
        results = search_excel_files(config_path, args.search, target)
        print(format_json({
            "keyword": args.search,
            "config_path": str(config_path),
            "matches": results,
            "total": len(results),
        }))
        return

    # 需要指定文件
    if not args.file:
        parser.error("--file is required for reading. Use --search to find files first.")

    filepath = resolve_file_path(config_path, args.file)

    # 列出 sheets
    if args.list_sheets:
        sheets = list_sheets(filepath)
        print(format_json({
            "file": str(filepath),
            "sheets": sheets,
            "total": len(sheets),
        }))
        return

    # 读取 sheet
    if not args.sheet:
        # 没有指定 sheet，先列出可用的 sheets
        sheets = list_sheets(filepath)
        print(f"No --sheet specified. Available sheets in {filepath.name}:", file=sys.stderr)
        for s in sheets:
            print(f"  - {s}", file=sys.stderr)
        parser.error("--sheet is required for reading data.")

    print(f"Reading: {filepath.name} -> {args.sheet}", file=sys.stderr)
    data = read_sheet(filepath, args.sheet, args.filter, args.limit)

    if args.format == "markdown":
        print(f"File: {filepath.name}, Sheet: {args.sheet}")
        print(f"Showing {data['total_shown']} of {data['total_matched']} matched rows\n")
        print(format_markdown_table(data["headers"], data["rows"]))
    else:
        print(format_json(data))


if __name__ == "__main__":
    main()
