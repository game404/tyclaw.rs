"""
Excel 配表 Diff 工具
从 Git 两个分支/commit 中提取同一个 Excel 文件，逐 sheet 逐行对比差异

用法示例:
  # 对比两个分支间所有变更的 xlsx 文件
  python tools/excel_diff.py --base origin/release/v2025.01.01 --head origin/release/v2025.01.15

  # 只对比指定文件
  python tools/excel_diff.py --base origin/release/v2025.01.01 --head origin/release/v2025.01.15 --file "角色配置.xlsx"

  # 指定 ID 列（默认使用第一列作为行 key）
  python tools/excel_diff.py --base origin/release/v2025.01.01 --head origin/release/v2025.01.15 --file "技能.xlsx" --key-col "SkillId"

  # 限制每个 sheet 的差异行数
  python tools/excel_diff.py --base origin/release/v2025.01.01 --head origin/release/v2025.01.15 --limit 30

  # markdown 格式输出
  python tools/excel_diff.py --base origin/release/v2025.01.01 --head origin/release/v2025.01.15 --format markdown

  # 审查今天的配表变更（热更场景）
  python tools/excel_diff.py --since today --format markdown

  # 审查过去2小时的变更
  python tools/excel_diff.py --hours 2 --format markdown

  # 审查从指定时间点开始的变更
  python tools/excel_diff.py --since "2026-03-01 07:31" --format markdown

输出结构 (JSON):
  {
    "base_ref": "origin/release/v2025.01.01",
    "head_ref": "origin/release/v2025.01.15",
    "files": [
      {
        "file": "角色配置.xlsx",
        "sheets": [
          {
            "sheet": "Sheet1",
            "added_rows": [...],
            "removed_rows": [...],
            "modified_rows": [{"key": "101", "changes": {"攻击力": {"old": 100, "new": 120}}}],
            "summary": "added: 2, removed: 0, modified: 3"
          }
        ]
      }
    ]
  }
"""

import argparse
import subprocess
import sys
import tempfile
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

from utils import load_config, format_json, get_repo_path


def get_config_path(config):
    """获取配表仓库路径"""
    config_path = get_repo_path(config, "config")
    if not config_path:
        print("Error: code.repos.config not configured in config.yaml", file=sys.stderr)
        sys.exit(1)
    path = Path(config_path)
    if not path.exists():
        print(f"Error: config path does not exist: {path}", file=sys.stderr)
        sys.exit(1)
    return path


def resolve_time_to_ref(repo_path, timestamp, head_ref):
    """将时间点解析为该时间之前最近的 git commit hash"""
    iso_ts = timestamp.strftime("%Y-%m-%dT%H:%M:%S")
    result = subprocess.run(
        ["git", "rev-list", f"--before={iso_ts}", "-1", head_ref],
        cwd=repo_path, capture_output=True, text=True
    )
    if result.returncode != 0:
        print(f"Error: git rev-list failed: {result.stderr}", file=sys.stderr)
        sys.exit(1)
    commit = result.stdout.strip()
    if not commit:
        return None
    return commit


def parse_since_value(since_str):
    """解析 --since 参数值，返回 datetime 对象"""
    if since_str.lower() == "today":
        return datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    if since_str.lower() == "yesterday":
        yesterday = datetime.now() - timedelta(days=1)
        return yesterday.replace(hour=0, minute=0, second=0, microsecond=0)
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(since_str, fmt)
        except ValueError:
            continue
    print(f"Error: invalid --since format: '{since_str}'. "
          f"Use 'today', 'yesterday', 'YYYY-MM-DD' or 'YYYY-MM-DD HH:MM'", file=sys.stderr)
    sys.exit(1)


def git_changed_xlsx_files(repo_path, base_ref, head_ref):
    """获取两个 ref 之间变更的 xlsx 文件列表"""
    result = subprocess.run(
        ["git", "diff", "--name-only", "--diff-filter=ACMR", base_ref, head_ref, "--", "*.xlsx"],
        cwd=repo_path, capture_output=True, text=True
    )
    if result.returncode != 0:
        print(f"Error: git diff failed: {result.stderr}", file=sys.stderr)
        sys.exit(1)
    files = [f.strip() for f in result.stdout.strip().split("\n") if f.strip()]
    return files


def git_deleted_xlsx_files(repo_path, base_ref, head_ref):
    """获取两个 ref 之间被删除的 xlsx 文件列表"""
    result = subprocess.run(
        ["git", "diff", "--name-only", "--diff-filter=D", base_ref, head_ref, "--", "*.xlsx"],
        cwd=repo_path, capture_output=True, text=True
    )
    if result.returncode != 0:
        return []
    return [f.strip() for f in result.stdout.strip().split("\n") if f.strip()]


def git_extract_file(repo_path, ref, filepath, dest_path):
    """从 git ref 中提取文件到本地临时路径，返回是否成功"""
    result = subprocess.run(
        ["git", "show", f"{ref}:{filepath}"],
        cwd=repo_path, capture_output=True
    )
    if result.returncode != 0:
        return False
    with open(dest_path, "wb") as f:
        f.write(result.stdout)
    return True


def read_sheet_data(wb, sheet_name):
    """读取 sheet 全部数据，返回 (headers, data_rows)"""
    ws = wb[sheet_name]
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(cell) if cell is not None else "" for cell in row]
        break
    if not headers:
        return headers, []

    data_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for i, val in enumerate(row):
            if i < len(headers):
                record[headers[i]] = val
        first_val = str(record.get(headers[0], "") or "")
        if first_val.startswith("##"):
            continue
        data_rows.append(record)

    return headers, data_rows


def normalize_value(val):
    """归一化单元格值用于比较"""
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return int(val)
    if isinstance(val, (str, int, float, bool)):
        return val
    # openpyxl ArrayFormula 等非基本类型 → 提取公式文本
    if hasattr(val, 'text'):
        return val.text or ""
    return str(val)


def values_equal(a, b):
    """比较两个单元格值是否相等"""
    return normalize_value(a) == normalize_value(b)


def _is_formula_row_shift(old_val, new_val):
    """检测两个值是否仅为公式行号偏移（结构相同只是数字不同）"""
    import re as _re
    old_s, new_s = str(old_val), str(new_val)
    if not old_s.startswith("=") or not new_s.startswith("="):
        return False
    return _re.sub(r"\d+", "N", old_s) == _re.sub(r"\d+", "N", new_s)


def _pick_key_column(headers, sample_rows, build_index_fn):
    """从 headers 中选出能产生最多有效索引的列作为 key 列"""
    if not headers:
        return ""
    best_col = headers[0]
    best_count = 0
    for col in headers:
        idx = build_index_fn(sample_rows, col)
        if len(idx) > best_count:
            best_count = len(idx)
            best_col = col
            if best_count >= len(sample_rows) * 0.8:
                break
    return best_col


def diff_sheet(base_headers, base_rows, head_headers, head_rows, key_col=None):
    """
    对比两个版本的 sheet 数据

    Args:
        base_headers: 旧版表头
        base_rows: 旧版数据行
        head_headers: 新版表头
        head_rows: 新版数据行
        key_col: 用于匹配行的列名，默认使用第一列
    Returns:
        dict: 差异结果
    """
    if not base_headers and not head_headers:
        return {"added_rows": [], "removed_rows": [], "modified_rows": [],
                "added_columns": [], "removed_columns": [], "summary": "empty sheet"}

    # 检测列变更
    base_col_set = set(base_headers)
    head_col_set = set(head_headers)
    added_columns = sorted(head_col_set - base_col_set)
    removed_columns = sorted(base_col_set - head_col_set)
    common_columns = [h for h in head_headers if h in base_col_set]

    # 建立 key -> row 的索引
    def build_index(rows, col):
        idx = {}
        for row in rows:
            key_val = str(normalize_value(row.get(col, "")))
            if key_val and key_val != "None" and key_val != "":
                idx[key_val] = row
        return idx

    # 确定 key 列：优先用指定列，否则自动选择能产生有效索引的列
    all_rows = base_rows or head_rows
    ref_headers = head_headers or base_headers
    if key_col and key_col in ref_headers:
        k = key_col
    else:
        k = _pick_key_column(ref_headers, all_rows, build_index)

    base_idx = build_index(base_rows, k)
    head_idx = build_index(head_rows, k)

    base_keys = set(base_idx.keys())
    head_keys = set(head_idx.keys())

    # 新增行
    added_keys = sorted(head_keys - base_keys, key=lambda x: (len(x), x))
    added_rows = []
    for ak in added_keys:
        row = head_idx[ak]
        compact = {h: normalize_value(row.get(h)) for h in head_headers
                   if normalize_value(row.get(h)) != ""}
        added_rows.append({"key": ak, "data": compact})

    # 删除行
    removed_keys = sorted(base_keys - head_keys, key=lambda x: (len(x), x))
    removed_rows = []
    for rk in removed_keys:
        row = base_idx[rk]
        compact = {h: normalize_value(row.get(h)) for h in base_headers
                   if normalize_value(row.get(h)) != ""}
        removed_rows.append({"key": rk, "data": compact})

    # 修改行（过滤掉纯公式行号偏移的变更）
    common_keys = base_keys & head_keys
    modified_rows = []
    formula_shift_rows = 0
    for ck in sorted(common_keys, key=lambda x: (len(x), x)):
        base_row = base_idx[ck]
        head_row = head_idx[ck]
        changes = {}
        for col in common_columns:
            old_val = normalize_value(base_row.get(col))
            new_val = normalize_value(head_row.get(col))
            if not values_equal(old_val, new_val):
                if _is_formula_row_shift(old_val, new_val):
                    continue
                changes[col] = {"old": old_val, "new": new_val}
        if changes:
            modified_rows.append({"key": ck, "changes": changes})
        elif any(not values_equal(normalize_value(base_row.get(c)),
                                  normalize_value(head_row.get(c)))
                 for c in common_columns):
            formula_shift_rows += 1

    summary_parts = []
    if added_rows:
        summary_parts.append(f"added: {len(added_rows)}")
    if removed_rows:
        summary_parts.append(f"removed: {len(removed_rows)}")
    if modified_rows:
        summary_parts.append(f"modified: {len(modified_rows)}")
    if formula_shift_rows:
        summary_parts.append(f"formula_shift_skipped: {formula_shift_rows}")
    if added_columns:
        summary_parts.append(f"new_cols: {added_columns}")
    if removed_columns:
        summary_parts.append(f"del_cols: {removed_columns}")

    # 公式偏移过滤后若无实际变更，返回 no changes
    has_changes = (added_rows or removed_rows or modified_rows
                   or added_columns or removed_columns)

    return {
        "key_column": k,
        "added_rows": added_rows,
        "removed_rows": removed_rows,
        "modified_rows": modified_rows,
        "added_columns": added_columns,
        "removed_columns": removed_columns,
        "summary": ", ".join(summary_parts) if has_changes else "no changes",
    }


def diff_excel_file(repo_path, base_ref, head_ref, filepath, key_col=None, limit=50):
    """
    对比单个 Excel 文件在两个 ref 之间的差异

    Returns:
        dict: 文件级别的差异结果
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        base_path = Path(tmpdir) / "base.xlsx"
        head_path = Path(tmpdir) / "head.xlsx"

        base_ok = git_extract_file(repo_path, base_ref, filepath, base_path)
        head_ok = git_extract_file(repo_path, head_ref, filepath, head_path)

        if not base_ok and not head_ok:
            return {"file": filepath, "error": "file not found in either ref"}

        if not base_ok:
            # 新增文件，列出所有 sheet
            try:
                wb = openpyxl.load_workbook(head_path, read_only=True)
                sheets_info = []
                for sn in wb.sheetnames:
                    headers, rows = read_sheet_data(wb, sn)
                    sheets_info.append({
                        "sheet": sn,
                        "status": "new_file",
                        "row_count": len(rows),
                        "headers": headers,
                    })
                wb.close()
                return {"file": filepath, "status": "added", "sheets": sheets_info}
            except Exception as e:
                return {"file": filepath, "status": "added", "error": str(e)}

        if not head_ok:
            return {"file": filepath, "status": "deleted"}

        # 两个版本都存在，逐 sheet 对比
        try:
            base_wb = openpyxl.load_workbook(base_path, read_only=True)
            head_wb = openpyxl.load_workbook(head_path, read_only=True)
        except Exception as e:
            return {"file": filepath, "error": f"failed to open workbook: {e}"}

        base_sheets = set(base_wb.sheetnames)
        head_sheets = set(head_wb.sheetnames)

        sheets_result = []

        # 新增的 sheet
        for sn in sorted(head_sheets - base_sheets):
            headers, rows = read_sheet_data(head_wb, sn)
            sheets_result.append({
                "sheet": sn,
                "status": "added",
                "row_count": len(rows),
                "headers": headers,
            })

        # 删除的 sheet
        for sn in sorted(base_sheets - head_sheets):
            sheets_result.append({"sheet": sn, "status": "deleted"})

        # 共同的 sheet，逐行 diff
        for sn in sorted(base_sheets & head_sheets):
            base_headers, base_rows = read_sheet_data(base_wb, sn)
            head_headers, head_rows = read_sheet_data(head_wb, sn)
            diff = diff_sheet(base_headers, base_rows, head_headers, head_rows, key_col)
            if diff["summary"] == "no changes":
                continue

            # 截断输出
            if limit:
                diff["added_rows"] = diff["added_rows"][:limit]
                diff["removed_rows"] = diff["removed_rows"][:limit]
                diff["modified_rows"] = diff["modified_rows"][:limit]

            sheets_result.append({"sheet": sn, "status": "modified", **diff})

        base_wb.close()
        head_wb.close()

        return {"file": filepath, "status": "modified", "sheets": sheets_result}


def _truncate(val, max_len=80):
    """截断过长的值，保留首尾"""
    s = str(val)
    if len(s) <= max_len:
        return s
    half = (max_len - 3) // 2
    return s[:half] + "..." + s[-half:]


def format_diff_markdown(result, *, skip_comment_cols=False):
    """将 diff 结果格式化为 markdown

    Args:
        skip_comment_cols: 为 True 时过滤 # 开头的备注列
    """
    lines = []
    for file_info in result.get("files", []):
        fname = file_info["file"]
        status = file_info.get("status", "")
        lines.append(f"## {fname} ({status})")

        if file_info.get("error"):
            lines.append(f"Error: {file_info['error']}\n")
            continue

        for sheet in file_info.get("sheets", []):
            sn = sheet["sheet"]
            s_status = sheet.get("status", "")
            lines.append(f"### Sheet: {sn} ({s_status})")

            if s_status in ("added", "deleted"):
                if "row_count" in sheet:
                    lines.append(f"Rows: {sheet['row_count']}")
                if "headers" in sheet:
                    lines.append(f"Columns: {', '.join(sheet['headers'][:20])}")
                lines.append("")
                continue

            summary = sheet.get("summary", "")
            lines.append(f"Key column: `{sheet.get('key_column', '?')}`  |  {summary}\n")

            if sheet.get("added_columns"):
                cols = [c for c in sheet['added_columns']
                        if not (skip_comment_cols and c.startswith('#'))]
                if cols:
                    lines.append(f"**New columns:** {', '.join(cols)}")
            if sheet.get("removed_columns"):
                cols = [c for c in sheet['removed_columns']
                        if not (skip_comment_cols and c.startswith('#'))]
                if cols:
                    lines.append(f"**Removed columns:** {', '.join(cols)}")

            if sheet.get("added_rows"):
                lines.append(f"**Added rows ({len(sheet['added_rows'])}):**")
                for r in sheet["added_rows"]:
                    preview = {k: _truncate(v) for k, v in list(r["data"].items())[:6]
                               if not (skip_comment_cols and k.startswith('#'))}
                    lines.append(f"- `{r['key']}`: {preview}")
                lines.append("")

            if sheet.get("removed_rows"):
                lines.append(f"**Removed rows ({len(sheet['removed_rows'])}):**")
                for r in sheet["removed_rows"]:
                    preview = {k: _truncate(v) for k, v in list(r["data"].items())[:6]
                               if not (skip_comment_cols and k.startswith('#'))}
                    lines.append(f"- `{r['key']}`: {preview}")
                lines.append("")

            if sheet.get("modified_rows"):
                rows_to_render = []
                for r in sheet["modified_rows"]:
                    changes = {col: cv for col, cv in r["changes"].items()
                               if not (skip_comment_cols and col.startswith('#'))}
                    if changes:
                        rows_to_render.append({"key": r["key"], "changes": changes})

                if rows_to_render:
                    lines.append(f"**Modified rows ({len(rows_to_render)}):**")
                    # 折叠重复变更：按变更签名分组
                    from collections import OrderedDict
                    groups: OrderedDict[tuple, list] = OrderedDict()
                    for r in rows_to_render:
                        sig = tuple(sorted(
                            (col, str(cv["old"]), str(cv["new"]))
                            for col, cv in r["changes"].items()
                        ))
                        groups.setdefault(sig, []).append(r)

                    for sig, members in groups.items():
                        first = members[0]
                        lines.append(f"- `{first['key']}`:")
                        for col, cv in first["changes"].items():
                            lines.append(f"  - {col}: `{_truncate(cv['old'])}` → `{_truncate(cv['new'])}`")
                        if len(members) > 1:
                            other_keys = [m["key"] for m in members[1:]]
                            if len(other_keys) <= 10:
                                keys_str = ", ".join(f"`{k}`" for k in other_keys)
                            else:
                                keys_str = ", ".join(f"`{k}`" for k in other_keys[:8])
                                keys_str += f" ... 等共 {len(other_keys)} 条"
                            lines.append(f"  - *(同上变更另有 {len(other_keys)} 行: {keys_str})*")
                    lines.append("")

        lines.append("")

    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(description="Excel config diff tool for BugHunter")
    parser.add_argument("--config", help="Path to config.yaml")
    parser.add_argument("--base", help="Base git ref (branch/tag/commit)")
    parser.add_argument("--head", help="Head git ref (branch/tag/commit, default: HEAD)")
    parser.add_argument("--since", help="Time-based base: 'today', 'yesterday', 'YYYY-MM-DD' or 'YYYY-MM-DD HH:MM'")
    parser.add_argument("--hours", type=float, help="Time-based base: changes in last N hours")
    parser.add_argument("--file", help="Diff only this file (relative to config repo root)")
    parser.add_argument("--key-col", help="Column name to use as row key (default: first column)")
    parser.add_argument("--limit", type=int, default=50, help="Max diff rows per sheet (default: 50)")
    parser.add_argument("--format", choices=["json", "markdown"], default="json", help="Output format")

    args = parser.parse_args()

    # --base / --since / --hours 三选一
    time_opts = sum(1 for x in [args.base, args.since, args.hours] if x is not None)
    if time_opts == 0:
        parser.error("must provide one of: --base, --since, --hours")
    if time_opts > 1:
        parser.error("--base, --since, --hours are mutually exclusive")

    config = load_config(args.config)
    repo_path = get_config_path(config)

    head_ref = args.head or "HEAD"

    # 时间范围解析
    if args.since is not None:
        ts = parse_since_value(args.since)
        print(f"Resolving --since '{args.since}' => {ts.strftime('%Y-%m-%d %H:%M')}...", file=sys.stderr)
        base_ref = resolve_time_to_ref(repo_path, ts, head_ref)
        if not base_ref:
            print(f"No commits found before {ts.strftime('%Y-%m-%d %H:%M')} on {head_ref}. "
                  f"All commits are within the specified range.", file=sys.stderr)
            # 用仓库最早的 commit 作为 base，这样能 diff 出全部内容
            result = subprocess.run(
                ["git", "rev-list", "--max-parents=0", head_ref],
                cwd=repo_path, capture_output=True, text=True
            )
            base_ref = result.stdout.strip().split("\n")[0]
        print(f"Base ref resolved to: {base_ref[:12]}", file=sys.stderr)
    elif args.hours is not None:
        ts = datetime.now() - timedelta(hours=args.hours)
        print(f"Resolving --hours {args.hours} => since {ts.strftime('%Y-%m-%d %H:%M')}...", file=sys.stderr)
        base_ref = resolve_time_to_ref(repo_path, ts, head_ref)
        if not base_ref:
            print(f"No commits found before {ts.strftime('%Y-%m-%d %H:%M')} on {head_ref}. "
                  f"All commits are within the specified range.", file=sys.stderr)
            result = subprocess.run(
                ["git", "rev-list", "--max-parents=0", head_ref],
                cwd=repo_path, capture_output=True, text=True
            )
            base_ref = result.stdout.strip().split("\n")[0]
        print(f"Base ref resolved to: {base_ref[:12]}", file=sys.stderr)
    else:
        base_ref = args.base

    if args.file:
        target_files = [args.file]
    else:
        print(f"Scanning changed xlsx files between {base_ref[:12]} and {head_ref}...", file=sys.stderr)
        target_files = git_changed_xlsx_files(repo_path, base_ref, head_ref)
        deleted_files = git_deleted_xlsx_files(repo_path, base_ref, head_ref)

        if not target_files and not deleted_files:
            print("No xlsx files changed.", file=sys.stderr)
            result = {"base_ref": base_ref, "head_ref": head_ref, "files": []}
            print(format_json(result))
            return

        print(f"Found {len(target_files)} changed, {len(deleted_files)} deleted xlsx files", file=sys.stderr)

    files_result = []

    for filepath in target_files:
        if filepath.startswith("~$"):
            continue
        print(f"Diffing: {filepath}", file=sys.stderr)
        diff = diff_excel_file(repo_path, base_ref, head_ref, filepath, args.key_col, args.limit)
        files_result.append(diff)

    if not args.file:
        for filepath in deleted_files:
            files_result.append({"file": filepath, "status": "deleted"})

    result = {
        "base_ref": base_ref,
        "head_ref": head_ref,
        "files": files_result,
    }

    if args.format == "markdown":
        print(format_diff_markdown(result))
    else:
        print(format_json(result))


if __name__ == "__main__":
    main()
